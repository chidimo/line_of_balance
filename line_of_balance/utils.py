"""Line of Balance curve plotter"""

import matplotlib.pyplot as plt
import matplotlib.path as mpath
import numpy as np
import openpyxl as OP
from .paths import OUT_PATH

# pylint: disable-msg=C0103
# pylint: disable-msg=E1101

def get_int_list(string_list):
    """Get a list of integers representing various list values"""
    values = re.split(r'[, \[ \]]', string_list)
    string_list = []
    for each in values:
        try:
            string_list.append(int(each))
        except ValueError:
            continue
    return string_list

def get_names_list(string_list):
    """Get list of activity names"""
    values = re.split(r'[, \[ \]]', string_list)
    values = [each.strip("'") for each in values if each != '']
    return values

def excel_table(headings, cell_values):
    """Table as excel file"""

    book = OP.Workbook()
    sheet = book.active
    sheet.title = "lineOfBalance"

    for idx, item in enumerate(headings):
        sheet.cell(row=1, column=idx+1, value=item)
    i = 2
    for each in cell_values:
        for idx, item in enumerate(each):
            sheet.cell(row=i, column=idx+1, value=item)
        i += 1
    book.save('output/line_of_balance.xlsx')

def pyplot_table(headings, cell_values):
    """Draw the line of balance table"""

    num_rows = len(cell_values[0])
    num_cols = len(cell_values)
    cell_width = 2.5
    cell_height = 0.3
    fig_width = cell_width * num_cols
    fig_height = cell_height * (num_rows + 1) # plus 1 for heading

    fig = plt.figure(figsize=(fig_width, fig_height))
    ax = fig.add_subplot(111)
    ax.axis('off')
    lob_table = plt.table(cellText=cell_values,
                          colLabels=headings,
                          loc='center',
                          bbox=[0, 0, 1, 0.5])

    lob_table.auto_set_font_size(False)
    lob_table.set_fontsize(5)
    lob_table.scale(2, 2)
    fig.savefig('line_of_balance_table.png', bbox_inches='tight', bbox_extra_artists=[lob_table])
    plt.show()

def plot_single_activity(coord1, coord2, coord3, coord4, ymin, ymax):
    """Generate plot from points"""

    _ = plt.figure(figsize=(10, 8), facecolor='lightblue', edgecolor='g')#fig
    label_x_position = (coord2 + coord3)//2
    label_y_position = (ymax + ymin)//2
    space = np.linspace(ymin, ymax, 5)
    plt.yticks(space)
    plt.ylabel("Number of units to produce")
    plt.xlabel("Project duration")
    plt.tight_layout()

    Path = mpath.Path
    path_data = [
        (Path.MOVETO, (coord1, ymin)),
        (Path.LINETO, (coord3, ymax)),
        (Path.LINETO, (coord4, ymax)),
        (Path.LINETO, (coord2, ymin)),
        (Path.CLOSEPOLY, (coord1, ymin))
    ]
    codes, verts = zip(*path_data)
    path = mpath.Path(verts, codes)

    # plot and fill control points and connecting lines
    x_vals, y_vals = zip(*path.vertices)
    plt.plot(x_vals, y_vals, 'k-')
    plt.fill(x_vals, y_vals, 'g')
    plt.text(label_x_position, label_y_position, "Plot")

    plt.plot([coord4, coord4], [ymin, ymax], 'r--')

    plt.show()

def plot_all_activities(set_of_points, ymin, ymax):
    """Generate plot from points"""

    project_duration = set_of_points[-1][4] + 1

    fig = plt.figure(facecolor='lightblue', edgecolor='g')
    axes = plt.gca()

    axes.set_xlim(0, 1.1*project_duration)
    axes.set_ylim(-0.1, 1.3*ymax)
    axes.spines['right'].set_visible(False)
    axes.spines['top'].set_visible(False)

    for each in set_of_points:
        name = each[0]
        coord1 = each[1]
        coord2 = each[2]
        coord3 = each[3]
        coord4 = each[4]

        Path = mpath.Path
        path_data = [
            (Path.MOVETO, (coord1, ymin)),
            (Path.LINETO, (coord3, ymax)),
            (Path.LINETO, (coord4, ymax)),
            (Path.LINETO, (coord2, ymin)),
            (Path.CLOSEPOLY, (coord1, ymin))
        ]

        codes, verts = zip(*path_data)
        path = mpath.Path(verts, codes)
        # calculate label position
        label_x_position = (coord2 + coord3)//2
        label_y_position = (ymax + ymin)//2

        # plot and fill control points and connecting lines
        x_vals, y_vals = zip(*path.vertices)
        plt.plot(x_vals, y_vals, 'b-', label=name)
        plt.fill(x_vals, y_vals, 'cyan')
        # add label
        plt.text(label_x_position, label_y_position, name)
        # draw dropdown
        plt.plot([coord4, coord4], [ymin, ymax], 'r--')

    plt.title("Line of Balance Curve")

    plt.annotate('Total project duration is\n{} days'.format(project_duration),
                 xy=(0, 0), xytext=(project_duration/2, 1.2*ymax),
                 horizontalalignment='center',
                 verticalalignment='top')

    space = np.linspace(0, 20, 5)
    plt.yticks(space)
    plt.ylabel("Number of units to produce")
    plt.xlabel("Project duration")

    plt.tight_layout()

    plt.savefig(OUT_PATH + ".pdf", facecolor=fig.get_facecolor(), dpi=100)
    plt.savefig(OUT_PATH + ".png", facecolor=fig.get_facecolor(), dpi=100)
    # plt.legend()
    plt.show()

def main():
    """Docstring"""
    pass
