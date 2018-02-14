"""Line of Balance curve plotter"""
import re
import matplotlib.pyplot as plt
import matplotlib.path as mpath
import numpy as np
import openpyxl as OP
try:
    import tkinter as tk
    from tkinter import filedialog
except:
    print("""tkinter could not be imported. You should check this error unless you're working on a cloud-based system.
          When you're done, press ENTER to continue'""")
    if input():
        pass

def open_save(mode="save"):
    """Return path to open or save file"""
    root = tk.Tk()
    root.withdraw()

    if mode == "open":
        return filedialog.askopenfilename()
    elif mode == "save":
        return filedialog.asksaveasfilename()

def get_int_list(string_list):
    """Get a list of integers representing various list values"""
    values = re.split(r'[, \[ \]]', string_list)
    string_list = []
    for each in values:
        try:
            string_list.append(int(each))
        except ValueError:
            continue
    return string_list

def get_names_list(string_list):
    """Get list of activity names"""
    values = re.split(r'[, \[ \]]', string_list)
    values = [each.strip("'") for each in values if each != '']
    return values

def parse_data():
    """Parse project activity data from text file"""
    try:
        with open(open_save(mode="open"), "r+") as input_handle:
            input_file_name = input_handle.readlines()
    except NameError:
        print("tk module not found")
        prompt = input("Please provide a file path\nType 'yes' to use default dataset\nENTER to exit")

        if not prompt:
            return
        elif prompt == 'yes':
            input_file_name = open('input.txt', 'r+')
        else:
            input_file_name = open(prompt, 'r+')

        input_file_name = input_file_name.readlines()

        activity_names = input_file_name[0].split('=')[1].strip()
        man_hours_per_unit = input_file_name[1].split('=')[1].strip()
        men_per_gang = input_file_name[2].split('=')[1].strip()

        buffer_time = int(input_file_name[3].split('=')[1].strip())
        productivity_rate = int(input_file_name[4].split('=')[1].strip())
        number_of_units_to_produce = int(input_file_name[5].split('=')[1].strip())
        hours_per_day = int(input_file_name[6].split('=')[1].strip())
        days_per_week = int(input_file_name[7].split('=')[1].strip())

        activity_names = get_names_list(activity_names)
        man_hours_per_unit = get_int_list(man_hours_per_unit)
        men_per_gang = get_int_list(men_per_gang)

        return activity_names, man_hours_per_unit, men_per_gang,\
        buffer_time, productivity_rate, number_of_units_to_produce,\
        hours_per_day, days_per_week
    return

def excel_table(headings, cell_values):
    """Table as excel file"""

    book = OP.Workbook()
    sheet = book.active
    sheet.title = "lineOfBalance"

    for idx, item in enumerate(headings):
        sheet.cell(row=1, column=idx+1, value=item)
    i = 2
    for each in cell_values:
        for idx, item in enumerate(each):
            sheet.cell(row=i, column=idx+1, value=item)
        i += 1

    try:
        name = filedialog.asksaveasfilename() + ".xlsx"
    except NameError:
        print("tk not found")
        name = 'line_ob_balance.xlsx'
    book.save(name)

def pyplot_table(headings, cell_values):
    """Draw the line of balance table in a plot"""

    num_rows = len(cell_values[0])
    num_cols = len(cell_values)
    cell_width = 2.5
    cell_height = 0.3
    fig_width = cell_width * num_cols
    fig_height = cell_height * (num_rows + 1) # plus 1 for heading

    fig = plt.figure(figsize=(fig_width, fig_height))
    ax = fig.add_subplot(111)
    ax.axis('off')
    lob_table = plt.table(cellText=cell_values,
                          colLabels=headings,
                          loc='center',
                          bbox=[0, 0, 1, 0.5])

    lob_table.auto_set_font_size(False)
    lob_table.set_fontsize(5)
    lob_table.scale(2, 2)
    fig.savefig('line_of_balance_table.png', bbox_inches='tight', bbox_extra_artists=[lob_table])
    plt.show()

def plot_single_activity(coord1, coord2, coord3, coord4, ymin, ymax):
    """Generate a single plot from a set of points"""

    _ = plt.figure(figsize=(10, 8), facecolor='lightblue', edgecolor='g')#fig
    label_x_position = (coord2 + coord3)//2
    label_y_position = (ymax + ymin)//2
    space = np.linspace(ymin, ymax, 5)
    plt.yticks(space)
    plt.ylabel("Number of units to produce")
    plt.xlabel("Project duration")
    plt.tight_layout()

    Path = mpath.Path
    path_data = [
        (Path.MOVETO, (coord1, ymin)),
        (Path.LINETO, (coord3, ymax)),
        (Path.LINETO, (coord4, ymax)),
        (Path.LINETO, (coord2, ymin)),
        (Path.CLOSEPOLY, (coord1, ymin))
    ]
    codes, verts = zip(*path_data)
    path = mpath.Path(verts, codes)

    # plot and fill control points and connecting lines
    x_vals, y_vals = zip(*path.vertices)
    plt.plot(x_vals, y_vals, 'k-')
    plt.fill(x_vals, y_vals, 'g')
    plt.text(label_x_position, label_y_position, "Plot")

    plt.plot([coord4, coord4], [ymin, ymax], 'r--')

    plt.show()

def plot_all_activities(set_of_points, ymin, ymax):
    """Generate multiple plot from a collection of points

    Parmaters
    ---------
    set_of_points : list
    """

    project_duration = set_of_points[-1][4] + 1

    fig = plt.figure(facecolor='lightblue', edgecolor='g')
    axes = plt.gca()

    axes.set_xlim(0, 1.1*project_duration)
    axes.set_ylim(-0.1, 1.3*ymax)
    axes.spines['right'].set_visible(False)
    axes.spines['top'].set_visible(False)

    for each in set_of_points:
        name = each[0]
        coord1 = each[1]
        coord2 = each[2]
        coord3 = each[3]
        coord4 = each[4]

        Path = mpath.Path
        path_data = [
            (Path.MOVETO, (coord1, ymin)),
            (Path.LINETO, (coord3, ymax)),
            (Path.LINETO, (coord4, ymax)),
            (Path.LINETO, (coord2, ymin)),
            (Path.CLOSEPOLY, (coord1, ymin))
        ]

        codes, verts = zip(*path_data)
        path = mpath.Path(verts, codes)
        # calculate label position
        label_x_position = (coord2 + coord3)//2
        label_y_position = (ymax + ymin)//2

        # plot and fill control points and connecting lines
        x_vals, y_vals = zip(*path.vertices)
        plt.plot(x_vals, y_vals, 'b-', label=name)
        plt.fill(x_vals, y_vals, 'cyan')
        # add label
        plt.text(label_x_position, label_y_position, name)
        # draw dropdown
        plt.plot([coord4, coord4], [ymin, ymax], 'r--')

    plt.title("Line of Balance Curve")

    plt.annotate('Total project duration is\n{} days'.format(project_duration),
                 xy=(0, 0), xytext=(project_duration/2, 1.2*ymax),
                 horizontalalignment='center',
                 verticalalignment='top')

    space = np.linspace(0, 20, 5)
    plt.yticks(space)
    plt.ylabel("Number of units to produce")
    plt.xlabel("Project duration")

    plt.tight_layout()
    # plt.legend()

    try:
        name = open_save()
    except NameError:
        print("tk not found")
        name = 'line_ob_balance'

    plt.savefig(name + ".pdf", facecolor=fig.get_facecolor(), dpi=100)
    plt.savefig(name + ".png", facecolor=fig.get_facecolor(), dpi=100)
    plt.show()

def illustrate_lob(coord1, coord2, coord3, coord4, ymin, ymax):
    """Generate plot from points"""

    fig = plt.figure(figsize=(10, 8), facecolor='lightblue', edgecolor='g')

    label_x_position = (coord2 + coord3)//2
    label_y_position = (ymax + ymin)//2
    space = np.linspace(ymin, ymax, 5)
    plt.yticks(space)
    plt.ylabel("Number of units to produce")
    plt.xlabel("Project duration")
    # plt.tight_layout()

    Path = mpath.Path
    path_data = [
        (Path.MOVETO, (coord1, ymin)),
        (Path.LINETO, (coord3, ymax)),
        (Path.LINETO, (coord4, ymax)),
        (Path.LINETO, (coord2, ymin)),
        (Path.CLOSEPOLY, (coord1, ymin))
    ]
    codes, verts = zip(*path_data)
    path = mpath.Path(verts, codes)

    # plot and fill control points and connecting lines
    x_vals, y_vals = zip(*path.vertices)
    plt.plot(x_vals, y_vals, 'b-')
    plt.fill(x_vals, y_vals, 'cyan')
    plt.text(label_x_position, label_y_position, "Activity")
    plt.plot([coord4, coord4], [ymin, ymax], 'r--')
    plt.title("Line of Balance Illustrator", color='k')

    # add annotations
    plt.annotate('Start on first section', xy=(coord1, ymin), xytext=(coord1, ymin+5),
                 arrowprops=dict(arrowstyle='->'), horizontalalignment='center')
    plt.annotate('End on first section', xy=(coord2, ymin), xytext=(coord2+5, ymin),
                 arrowprops=dict(arrowstyle='->'), verticalalignment='right')
    plt.annotate('Start on last section', xy=(coord3, ymax), xytext=(coord3, ymax+4),
                 arrowprops=dict(arrowstyle='->'), horizontalalignment='center')
    plt.annotate('End on last section', xy=(coord4, ymax), xytext=(coord4, ymax+2),
                 arrowprops=dict(arrowstyle='->'), horizontalalignment='center')

    axes = plt.gca()
    axes.title.set_color('red')

    axes.set_xlim(0, 1.1*coord4)
    axes.set_ylim(-0.1, 1.3*ymax)
    axes.spines['right'].set_visible(False)
    axes.spines['top'].set_visible(False)

    try:
        name = open_save()
    except NameError:
        print("tk not found")
        name = 'line_ob_balance'

    plt.savefig(name + ".pdf", facecolor=fig.get_facecolor(), dpi=100)
    plt.savefig(name + ".png", facecolor=fig.get_facecolor(), dpi=100)
    plt.show()

def default_lob():
    """Plot curve with default values for 5 activities"""
    activity_names = ['A', 'B', 'C', 'D', 'E']
    man_hours_per_unit = [100, 350, 60, 200, 150]
    men_per_gang = [4, 6, 2, 5, 8]
    buffer_time = 5
    productivity_rate = 3
    number_of_units_to_produce = 20
    hours_per_day = 8
    days_per_week = 5

    line_object = LineOfBalance(activity_names,
                                man_hours_per_unit,
                                men_per_gang,
                                buffer_time,
                                productivity_rate,
                                number_of_units_to_produce,
                                hours_per_day,
                                days_per_week)
    line_object.generate_curve()
    line_object.generate_excel_table()
    return line_object

class LineOfBalance:
    """LINE OF BALANCE
    Parameters
    -----------
    activity names : list
    man_hours_per_gang : list
    men_per_gang : list
    buffer_time : int
    productivity_rate : int
    number_of_units_to_produce : int
    hours_per_day : int
    days_per_week : int
    ymin : int"""

    def __init__(self, activity_names, man_hours_per_unit, men_per_gang, buffer_time, productivity_rate, number_of_units_to_produce, hours_per_day, days_per_week, ymin=0):

        self.activity_names = activity_names
        self.man_hours_per_unit = man_hours_per_unit
        self.men_per_gang = men_per_gang
        self.buffer_time = buffer_time
        self.productivity_rate = productivity_rate
        self.number_of_units_to_produce = number_of_units_to_produce
        self.hours_per_day = hours_per_day
        self.days_per_week = days_per_week
        self.number_of_activities = len(self.activity_names)
        self.ymax = self.number_of_units_to_produce
        self.ymin = ymin

    def theoretical_gang_size(self):
        """Compute theoretical gang sizes"""
        theo_g_size = [(self.productivity_rate*each) /
                       (self.hours_per_day*self.days_per_week)
                       for each in self.man_hours_per_unit]
        return [round(x, 2) for x in theo_g_size]

    def actual_gang_size(self):
        """Compute actual gang sizes"""
        actual_gang_size = [2 * x for x in self.men_per_gang]

        for idx, item in enumerate(actual_gang_size):
            old_gang_size = item

            for j in range(3, 6):
                new_gang_size = j * self.men_per_gang[idx]

                if abs(old_gang_size - self.theoretical_gang_size()[idx]) >=\
                abs(new_gang_size - self.theoretical_gang_size()[idx]):
                    old_gang_size = new_gang_size

            actual_gang_size[idx] = old_gang_size
        return list([round(x, 2) for x in actual_gang_size])

    def actual_output_rate(self):
        """Compute actual output rates"""
        act_out = [self.productivity_rate * (x/y)
                   for x, y in zip(self.actual_gang_size(),
                                   self.theoretical_gang_size())]
        return list([round(x, 2) for x in act_out])

    def activity_duration_per_unit(self):
        """Compute actual duration per unit"""
        adpu = [mhpu / (mpg * self.hours_per_day)
                for  mhpu, mpg in zip(self.man_hours_per_unit, self.men_per_gang)]
        return [round(x, 2) for x in adpu]

    def start_first_to_start_last(self):
        """Docstring"""
        sftl = [(self.number_of_units_to_produce - 1) * self.days_per_week /
                each for each in self.actual_output_rate()]
        return [round(x, 2) for x in sftl]

    def compute_plot_points(self):
        """Compute four last columns concurrently. First section start and end"""
        start_first = [1]
        end_first = [self.activity_duration_per_unit()[0]]

        start_last = [self.start_first_to_start_last()[0]]
        end_last = [start_last[0] + end_first[0]]

        for i in range(1, self.number_of_activities):

            if self.actual_output_rate()[i] < self.actual_output_rate()[i-1]:
                # place buffer at bottom
                start_first.insert(i, end_first[i-1] + self.buffer_time)
                end_first.insert(i, start_first[i] + self.activity_duration_per_unit()[i])

                start_last.insert(i, start_first[i] +\
                self.start_first_to_start_last()[i])
                end_last.insert(i, start_last[i] + self.activity_duration_per_unit()[i])

            elif self.actual_output_rate()[i] > self.actual_output_rate()[i-1]:
                # place buffer at bottom
                start_last.insert(i, end_last[i-1] + self.buffer_time)
                end_last.insert(i, start_last[i] + self.activity_duration_per_unit()[i])

                start_first.insert(i, start_last[i] - \
                self.start_first_to_start_last()[i])
                end_first.insert(i, start_first[i] + self.activity_duration_per_unit()[i])

        return [round(x, 2) for x in start_first],\
               [round(x, 2) for x in end_first],\
               [round(x, 2) for x in start_last],\
               [round(x, 2) for x in end_last]

    def start_on_first_section(self):
        """Start on first section for activity"""
        return self.compute_plot_points()[0]

    def end_on_first_section(self):
        """End on first section for activity"""
        return self.compute_plot_points()[1]

    def start_on_last_section(self):
        """Start on last section for activity"""
        return self.compute_plot_points()[2]

    def end_on_last_section(self):
        """End on last section for activity"""
        return self.compute_plot_points()[3]

    def arrange_values(self):
        """Arrange the values in tabular format"""
        values = [
            [a, b, c, d, e, f, g, h, i, j, k, l]
                  for a, b, c, d, e, f, g, h, i, j, k, l in
                  zip(self.activity_names,
                      self.man_hours_per_unit,
                      self.men_per_gang,
                      self.theoretical_gang_size(),
                      self.actual_gang_size(),
                      self.actual_output_rate(),
                      self.activity_duration_per_unit(),
                      self.start_first_to_start_last(),
                      self.start_on_first_section(),
                      self.end_on_first_section(),
                      self.start_on_last_section(),
                      self.end_on_last_section())
                ]
        return values

    def column_headings(self): # add **kwargs to take additional columns
        """Column headings"""
        columns = ("Activity",
                   "Man hours per unit",
                   "Men per gang",
                   "Theoretical gang size",
                   "Actual gang size",
                   "Actual output rate",
                   "Activity duration per unit",
                   "Start on first section to start on last section",
                   "Start on first section",
                   "End on first section",
                   "Start on last section",
                   "End on last section")
        return columns

    def plot_points_with_labels(self):
        """Generate points to be used for the plot.
        Points used in plotting are:

        Start on first section, End on first section,
        Start on last section, End on last section
        """
        return [(a, b, c, d, e) for a, b, c, d, e in zip(self.activity_names,
                                                         self.start_on_first_section(),
                                                         self.end_on_first_section(),
                                                         self.start_on_last_section(),
                                                         self.end_on_last_section())]

    def generate_curve(self):
        """Generate the line of balance curve"""
        plot_all_activities(self.plot_points_with_labels(), self.ymin, self.ymax)

    def project_duration(self):
        """Total project duration"""
        return self.plot_points_with_labels()[-1][4]+1

    def generate_excel_table(self):
        """Create table from object"""
        heads = self.column_headings()
        vals = self.arrange_values()
        excel_table(heads, vals)

def plot_lob_curve():
    """Generate curve for user input
    Returns
    --------
    Line of Balance object
    """
    data = parse_data()

    line_object = LineOfBalance(data[0],
                                data[1],
                                data[2],
                                data[3],
                                data[4],
                                data[5],
                                data[6],
                                data[7])
    line_object.generate_excel_table()
    line_object.generate_curve()
    return line_object

if __name__ == "__main__":
    print("Working")
    illustrate_lob(5, 15, 35, 45, 0, 20)