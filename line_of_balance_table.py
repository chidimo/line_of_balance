"""Line of Balance table generator

author: Chidi Orji
email: orjichidi95@gmail.com
github: https://github.com/Parousiaic
license: BSD
Please feel free to use and modify this, but keep the above information. Thanks!
I bear no responsibility for anything you might break while using this script.
"""

import matplotlib.pyplot as plt
import openpyxl as OP

# pylint: disable-msg=C0103
# pylint: disable-msg=E1101

def excel_table(headings, cell_values):
    """Table as excel file"""

    book = OP.Workbook()
    sheet = book.active
    sheet.title = "lineOfBalance"

    for idx, item in enumerate(headings):
        sheet.cell(row=1, column=idx+1, value=item)
    i = 2
    for each in cell_values:
        for idx, item in enumerate(each):
            sheet.cell(row=i, column=idx+1, value=item)
        i += 1
    book.save('output/line_of_balance.xlsx')

def pyplot_table(headings, cell_values):
    """Draw the line of balance table"""

    num_rows = len(cell_values[0])
    num_cols = len(cell_values)
    cell_width = 2.5
    cell_height = 0.3
    fig_width = cell_width * num_cols
    fig_height = cell_height * (num_rows + 1) # plus 1 for heading

    fig = plt.figure(figsize=(fig_width, fig_height))
    ax = fig.add_subplot(111)
    ax.axis('off')
    lob_table = plt.table(cellText=cell_values,
                          colLabels=headings,
                          loc='center',
                          bbox=[0, 0, 1, 0.5])

    lob_table.auto_set_font_size(False)
    lob_table.set_fontsize(5)
    lob_table.scale(2, 2)
    fig.savefig('line_of_balance_table.png', bbox_inches='tight', bbox_extra_artists=[lob_table])
    plt.show()

def main():
    """Nothing to put here for now"""
    pass

if __name__ == "__main__":
    main()
    